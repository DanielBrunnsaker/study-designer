from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def make_excel_report(
    seq_df: pd.DataFrame,
    explanation: str,
    run_config: dict,
) -> bytes:
    """
    Build a two-sheet Excel workbook:
      Sheet 1 — Run Sequence (injection order table)
      Sheet 2 — Study Report (LLM explanation + config parameters)
    """
    wb = Workbook()

    # ── Sheet 1: Run Sequence ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Run Sequence"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    type_colors = {
        "Sample": "C6EFCE",
        "QC": "FFEB9C",
        "Blank": "BDD7EE",
        "Wash": "E2EFDA",
        "Standard": "FCE4D6",
    }

    rows = list(dataframe_to_rows(seq_df.reset_index(), index=False, header=True))
    for r_idx, row in enumerate(rows, start=1):
        ws1.append(row)
        if r_idx == 1:
            for cell in ws1[r_idx]:
                cell.font = header_font
                cell.fill = header_fill
        else:
            sample_type = str(row[2]) if len(row) > 2 else ""
            color = type_colors.get(sample_type)
            if color:
                for cell in ws1[r_idx]:
                    cell.fill = PatternFill("solid", fgColor=color)

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    # ── Sheet 2: Study Report ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Study Report")

    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=11)

    ws2.append(["Metabolomics Study Design Report"])
    ws2["A1"].font = title_font
    ws2.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws2.append([])

    ws2.append(["Design Explanation"])
    ws2[f"A{ws2.max_row}"].font = section_font
    for line in explanation.split("\n"):
        ws2.append([line])

    ws2.append([])
    ws2.append(["Run Configuration"])
    ws2[f"A{ws2.max_row}"].font = section_font
    ws2.append(["Parameter", "Value"])
    header_row = ws2.max_row
    for cell in ws2[header_row]:
        cell.font = header_font
        cell.fill = header_fill

    param_labels = {
        "qc_frequency": "QC every N samples",
        "qc_at_start": "QCs at start",
        "qc_at_end": "QCs at end",
        "blank_at_start": "Blank at start",
        "blank_at_end": "Blank at end",
        "blank_frequency": "Periodic blank every N (0=none)",
        "wash_after_blank": "Wash after blank",
        "randomize_samples": "Randomize samples",
        "stratify_by_group": "Stratify by group",
        "block_by_batch": "Block by batch",
        "randomize_seed": "Random seed",
    }
    for key, label in param_labels.items():
        if key in run_config:
            ws2.append([label, str(run_config[key])])

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
